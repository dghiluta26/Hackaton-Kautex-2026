"""Export service: turns the current database contents into CSV/Excel files.

Used by the Reports page's "Export Data" tab.
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlmodel import select

from database.connection import get_session
from models.allocation import Allocation
from models.cost_item import CostItem
from models.employee import Employee
from models.topic import Topic
from services.cost_service import get_all_topic_cost_breakdowns, get_global_cost_totals
from utils.calculations import calculate_employee_topic_cost


def export_employees_csv() -> bytes:
    with get_session() as session:
        employees = session.exec(select(Employee)).all()

    df = pd.DataFrame(
        [
            {
                "ID": e.id,
                "Name": e.name,
                "Team": e.team_id if e.team_id is not None else "Unassigned",
                "Department": e.department_id if e.department_id is not None else "Unassigned",
                "Location": e.location_id if e.location_id is not None else "Unassigned",
                "Available Hours/Year": e.available_hours_per_year,
                "Hourly Rate": e.hourly_rate,
                "Status": e.status,
                "Manager": e.manager,
            }
            for e in employees
        ]
    )
    return df.to_csv(index=False).encode()


def export_topics_csv() -> bytes:
    with get_session() as session:
        topics = session.exec(select(Topic)).all()

    df = pd.DataFrame(
        [
            {
                "ID": t.id,
                "Name": t.name,
                "Category": t.category,
                "Area": t.area,
                "Status": t.status,
                "Business Justification": t.business_justification,
            }
            for t in topics
        ]
    )
    return df.to_csv(index=False).encode()


def export_allocations_csv() -> bytes:
    with get_session() as session:
        statement = select(Allocation, Employee, Topic).where(
            (Allocation.employee_id == Employee.id) & (Allocation.topic_id == Topic.id)
        )
        results = session.exec(statement).all()

    df = pd.DataFrame(
        [
            {
                "Employee": emp.name,
                "Topic": topic.name,
                "Allocation %": f"{alloc.allocation_percentage:.1f}%",
                "Allocated Hours": f"{emp.available_hours_per_year * (alloc.allocation_percentage / 100):.0f}",
                "Cost": f"${calculate_employee_topic_cost(emp.available_hours_per_year, emp.hourly_rate, alloc.allocation_percentage):,.2f}",
                "Comment": alloc.comment,
            }
            for alloc, emp, topic in results
        ]
    )
    return df.to_csv(index=False).encode()


def export_cost_report_excel() -> bytes:
    """Comprehensive cost report: Summary, Topics and Allocations sheets."""
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    breakdowns = get_all_topic_cost_breakdowns()
    metrics = get_global_cost_totals(breakdowns)

    ws_summary["A1"] = "Kautex Resource & Cost Planning - Summary Report"
    ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary["A3"] = "Total Cost:"
    ws_summary["B3"] = metrics["total_cost"]
    ws_summary["A4"] = "Total Headcount:"
    ws_summary["B4"] = metrics["total_headcount"]
    ws_summary["A5"] = "Avg Cost/Employee:"
    ws_summary["B5"] = metrics["average_cost_per_employee"]

    with get_session() as session:
        topics = session.exec(select(Topic)).all()
        allocations = session.exec(
            select(Allocation, Employee, Topic).where(
                (Allocation.employee_id == Employee.id) & (Allocation.topic_id == Topic.id)
            )
        ).all()

    ws_topics = wb.create_sheet("Topics")
    ws_topics.append(["Topic", "Category", "Internal Cost", "External Tooling", "Testing", "Recovery", "Total Cost"])
    for topic in topics:
        cost = breakdowns.get(topic.id, {"internal_personnel": 0, "external_tooling": 0, "testing": 0, "recovery": 0, "total": 0})
        ws_topics.append(
            [topic.name, topic.category, cost["internal_personnel"], cost["external_tooling"], cost["testing"], cost["recovery"], cost["total"]]
        )

    ws_alloc = wb.create_sheet("Allocations")
    ws_alloc.append(["Employee", "Topic", "Allocation %", "Allocated Hours", "Cost", "Comment"])
    for alloc, emp, topic in allocations:
        hours = emp.available_hours_per_year * (alloc.allocation_percentage / 100)
        cost = calculate_employee_topic_cost(emp.available_hours_per_year, emp.hourly_rate, alloc.allocation_percentage)
        ws_alloc.append([emp.name, topic.name, f"{alloc.allocation_percentage:.1f}%", f"{hours:.0f}", f"${cost:,.2f}", alloc.comment])

    for ws in [ws_summary, ws_topics, ws_alloc]:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_allocation_matrix_excel() -> bytes:
    """Allocation matrix as a color-coded Excel sheet (employees x topics)."""
    with get_session() as session:
        employees = session.exec(select(Employee)).all()
        topics = session.exec(select(Topic)).all()
        allocations = session.exec(select(Allocation)).all()

    alloc_lookup = {(a.employee_id, a.topic_id): a.allocation_percentage for a in allocations}

    matrix_data = []
    for emp in employees:
        row = {"Employee": emp.name, "Hourly Rate": emp.hourly_rate}
        for topic in topics:
            row[topic.name] = alloc_lookup.get((emp.id, topic.id), 0.0)
        row["Total %"] = sum(row[t.name] for t in topics)
        matrix_data.append(row)

    df = pd.DataFrame(matrix_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Allocation Matrix"

    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(r, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="00A6DF", end_color="00A6DF", fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
